"""Excel (.xlsx) renderer — real Shape objects + swimlane background.

Design
------
* Sheet ``フロー図``: swimlane bands as cell backgrounds + nodes/edges as
  real Excel Shape objects (drawing). The user can drag, resize and retext
  shapes; the sheet as a whole is overwritten on regeneration.
* Sheet ``注記``: free-form notes. Preserved across regeneration — if the
  target file already exists, its ``注記`` sheet is read and carried over.

Layout
------
All rank / collision / back-edge classification lives in
:mod:`jsox_flow.layout`. This module consumes a :class:`LayoutResult` and
translates its px coordinates into Excel's EMU / cell-anchor space.

Edges are emitted as explicit orthogonal paths. That keeps the PDF/PNG
rendering stable across Excel and LibreOffice, and avoids connector
auto-routing that can make back edges appear to attach to the wrong node.

Implementation
--------------
1. Build the base workbook with openpyxl (cells + formatting only).
2. Save to an in-memory buffer.
3. Post-process the xlsx ZIP: inject ``xl/drawings/drawing1.xml`` with
   ``<xdr:sp>`` for nodes and ``<xdr:cxnSp>`` for edges, and wire up the
   relationship (``sheet1.xml.rels``), content-type, and ``<drawing>``
   reference in ``sheet1.xml``.
"""
from __future__ import annotations

import io
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Union
from xml.sax.saxutils import escape as xml_escape

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..layout import LayoutOptions, LayoutResult, compute_layout
from ..model import Flow, Node


# --- EMU conversion ---------------------------------------------------------

EMU_PER_PX = 9525
EMU_PER_PT = 12700
PX_PER_CHAR = 7.0
PX_PER_PT = 96 / 72


# --- sheet structure --------------------------------------------------------

SHEET_FLOW = "フロー図"
SHEET_NOTES = "注記"

TITLE_ROW = 1
WARNING_ROW = 2

# Horizontal: lane-header in col A, step cells from col B. Row 3 is a
# back-edge channel so U-shaped connectors have clean space to route.
H_BACK_CHANNEL_ROW = 3
H_FIRST_LANE_ROW = 4
H_LANE_HEADER_COL = 1
H_FIRST_STEP_COL = 2

# Vertical: lane-header in row 3, step cells from row 4. Col A is the
# back-edge channel so U-shaped connectors have clean space to route.
V_LANE_HEADER_ROW = 3
V_FIRST_LANE_COL = 2
V_FIRST_STEP_ROW = 4

ROW_TITLE_HEIGHT_PT = 28
ROW_WARNING_HEIGHT_PT = 32
ROW_BACK_CHANNEL_HEIGHT_PT = 32
COL_BACK_CHANNEL_WIDTH_CHARS = 12

H_HEADER_WIDTH_CHARS = 14


# --- styling ----------------------------------------------------------------

NODE_PRESET = {
    "start":      "flowChartTerminator",
    "end":        "flowChartTerminator",
    "task":       "flowChartProcess",
    "decision":   "flowChartDecision",
    "subprocess": "flowChartPredefinedProcess",
    "document":   "flowChartDocument",
}
NODE_FILL = {
    "start":      "FFF3CD",
    "end":        "FFF3CD",
    "task":       "FFFFFF",
    "decision":   "D1ECF1",
    "subprocess": "E2E3E5",
    "document":   "FEFEFE",
}
LANE_HEADER_FILL = "D4DFF0"
LANE_BODY_FILLS = ("F7F9FC", "EEF2F8")
TITLE_FILL = "2F5597"
WARNING_FILL = "FFF2CC"
EDGE_COLOR = "333333"
LABEL_COLOR = "C0392B"

LABEL_W_PX = 50
LABEL_H_PX = 18
CONNECTOR_PAD_PX = 6
LOCAL_BACK_CHANNEL_GAP_PX = 18


# --- connection-point indices ----------------------------------------------
# per the OOXML preset definitions (<cxnLst> order) for all the flowChart*
# presets we emit: TOP, LEFT, BOTTOM, RIGHT.
SITE_TOP = 0
SITE_LEFT = 1
SITE_BOTTOM = 2
SITE_RIGHT = 3


# --- grid record -----------------------------------------------------------

@dataclass
class _Grid:
    """The Excel cell grid used as a visual backing for the swimlanes."""
    vertical: bool
    n_lanes: int
    n_ranks: int
    col_widths_chars: List[float]   # 1-indexed; index 0 unused
    row_heights_pt: List[float]
    col_edges_emu: List[int]        # cumulative EMU at each column edge
    row_edges_emu: List[int]

    def body_origin_emu(self) -> Tuple[int, int]:
        """(x, y) EMU offset from top-left of sheet to top-left of body."""
        if self.vertical:
            return (self.col_edges_emu[V_FIRST_LANE_COL - 1],
                    self.row_edges_emu[V_FIRST_STEP_ROW - 1])
        return (self.col_edges_emu[H_FIRST_STEP_COL - 1],
                self.row_edges_emu[H_FIRST_LANE_ROW - 1])

    def back_channel_center_emu(self) -> int:
        """Center of the back-edge channel on the perpendicular axis."""
        if self.vertical:
            # col A is the back-edge channel
            return (self.col_edges_emu[0] + self.col_edges_emu[1]) // 2
        # row 3 is the back-edge channel
        return (self.row_edges_emu[H_BACK_CHANNEL_ROW - 1]
                + self.row_edges_emu[H_BACK_CHANNEL_ROW]) // 2

    def last_col(self) -> int:
        if self.vertical:
            return V_FIRST_LANE_COL + self.n_lanes - 1
        return H_FIRST_STEP_COL + self.n_ranks - 1

    def last_row(self) -> int:
        if self.vertical:
            return V_FIRST_STEP_ROW + self.n_ranks - 1
        return H_FIRST_LANE_ROW + self.n_lanes - 1


# --- public API -------------------------------------------------------------

def save_xlsx(
    flow: Flow,
    path: Union[str, Path],
    *,
    vertical: bool = False,
    layout: Optional[LayoutResult] = None,
    return_layout: bool = False,
) -> Union[Path, Tuple[Path, LayoutResult]]:
    """Render ``flow`` to an .xlsx file.

    Parameters
    ----------
    layout:
        Optional pre-computed layout. If omitted, one is computed with
        the default :class:`LayoutOptions` for the requested orientation.
    return_layout:
        If True, return a ``(path, layout)`` tuple so the caller (agent)
        can inspect metrics / warnings without recomputing.
    """
    target = Path(path)
    preserved_notes = _extract_notes(target)

    orientation = "vertical" if vertical else "horizontal"
    if layout is None:
        layout = compute_layout(flow, LayoutOptions(orientation=orientation))
    elif layout.orientation != orientation:
        raise ValueError(
            f"layout.orientation={layout.orientation!r} but "
            f"vertical={vertical!r} requires {orientation!r}"
        )

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_FLOW

    grid = _make_grid(flow, layout)
    _build_flow_sheet(ws, flow, layout, grid)

    notes_ws = wb.create_sheet(SHEET_NOTES)
    _build_notes_sheet(notes_ws, preserved_notes)

    buf = io.BytesIO()
    wb.save(buf)

    drawing_xml = _build_drawing_xml(flow, layout, grid)
    final_bytes = _inject_drawing(buf.getvalue(), drawing_xml)
    target.write_bytes(final_bytes)

    if return_layout:
        return target, layout
    return target


# --- grid construction ------------------------------------------------------

def _make_grid(flow: Flow, layout: LayoutResult) -> _Grid:
    vertical = layout.orientation == "vertical"
    n_lanes = len(layout.lane_order)
    n_ranks = layout.metrics.n_ranks or 1

    # Convert px grid sizes to Excel cell sizes.
    step_size_px = _step_size_px(layout)
    lane_size_px = _lane_size_px(layout)

    step_width_chars = step_size_px / PX_PER_CHAR
    lane_height_pt = lane_size_px / PX_PER_PT
    lane_width_chars = lane_size_px / PX_PER_CHAR
    step_height_pt = step_size_px / PX_PER_PT

    if vertical:
        col_widths_chars = [0.0, COL_BACK_CHANNEL_WIDTH_CHARS]
        col_widths_chars.extend([lane_width_chars] * n_lanes)
        row_heights_pt = [
            0.0,
            ROW_TITLE_HEIGHT_PT,
            ROW_WARNING_HEIGHT_PT,
            lane_size_px / PX_PER_PT / 2,   # lane header row
        ]
        row_heights_pt.extend([step_height_pt] * n_ranks)
    else:
        col_widths_chars = [0.0, H_HEADER_WIDTH_CHARS]
        col_widths_chars.extend([step_width_chars] * n_ranks)
        row_heights_pt = [
            0.0,
            ROW_TITLE_HEIGHT_PT,
            ROW_WARNING_HEIGHT_PT,
            ROW_BACK_CHANNEL_HEIGHT_PT,
        ]
        row_heights_pt.extend([lane_height_pt] * n_lanes)

    col_edges = [0]
    for w in col_widths_chars[1:]:
        col_edges.append(col_edges[-1] + int(w * PX_PER_CHAR * EMU_PER_PX))
    row_edges = [0]
    for h in row_heights_pt[1:]:
        row_edges.append(row_edges[-1] + int(h * EMU_PER_PT))

    return _Grid(
        vertical=vertical,
        n_lanes=n_lanes,
        n_ranks=n_ranks,
        col_widths_chars=col_widths_chars,
        row_heights_pt=row_heights_pt,
        col_edges_emu=col_edges,
        row_edges_emu=row_edges,
    )


def _step_size_px(layout: LayoutResult) -> int:
    if layout.metrics.n_ranks <= 1:
        return layout.canvas_width if layout.orientation == "horizontal" else layout.canvas_height
    total = layout.canvas_width if layout.orientation == "horizontal" else layout.canvas_height
    return total // layout.metrics.n_ranks


def _lane_size_px(layout: LayoutResult) -> int:
    if layout.metrics.n_lanes <= 0:
        return 120
    total = layout.canvas_height if layout.orientation == "horizontal" else layout.canvas_width
    return total // layout.metrics.n_lanes


# --- notes sheet preservation ----------------------------------------------

def _extract_notes(path: Path) -> List[List]:
    if not path.exists():
        return []
    try:
        existing = load_workbook(path)
    except Exception:
        return []
    if SHEET_NOTES not in existing.sheetnames:
        return []
    ws = existing[SHEET_NOTES]
    rows: List[List] = []
    for row in ws.iter_rows(values_only=True):
        if any(cell not in (None, "") for cell in row):
            rows.append(list(row))
    return rows


def _build_notes_sheet(ws, preserved: List[List]) -> None:
    ws.column_dimensions["A"].width = 90
    ws.column_dimensions["B"].width = 20
    if not preserved:
        ws["A1"] = "このシートは手書き専用です。再生成時も内容は保持されます。"
        ws["A1"].font = Font(italic=True, color="808080")
        return
    for r, row in enumerate(preserved, start=1):
        for c, val in enumerate(row, start=1):
            ws.cell(r, c, val)


# --- flow sheet (cells) -----------------------------------------------------

def _build_flow_sheet(
    ws, flow: Flow, layout: LayoutResult, grid: _Grid,
) -> None:
    for c in range(1, len(grid.col_widths_chars)):
        ws.column_dimensions[get_column_letter(c)].width = grid.col_widths_chars[c]
    for r in range(1, len(grid.row_heights_pt)):
        ws.row_dimensions[r].height = grid.row_heights_pt[r]

    last_col = grid.last_col()
    last_row = grid.last_row()

    _write_title(ws, last_col)
    _write_warning(ws, last_col, layout.warnings)
    _write_lane_bands(ws, flow, layout, grid, last_row, last_col)
    _configure_printing(ws, grid, last_row, last_col)

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B4"


def _configure_printing(
    ws, grid: _Grid, last_row: int, last_col: int,
) -> None:
    """Set up print area so the PDF / printout is readable.

    Horizontal flowcharts easily exceed A4 landscape; forcing them onto
    a single page makes text microscopic. We use A3 landscape for those
    (which fits most J-SOX flows at a readable size) and only then turn
    on fit-to-width.
    """
    if grid.vertical:
        ws.page_setup.orientation = "portrait"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
    else:
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # 0 = unlimited height in pages
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4
    last_letter = get_column_letter(last_col)
    ws.print_area = f"A1:{last_letter}{last_row}"


def _write_title(ws, last_col: int) -> None:
    ws.cell(TITLE_ROW, 1, "業務フロー図")
    ws.merge_cells(
        start_row=TITLE_ROW, start_column=1,
        end_row=TITLE_ROW, end_column=last_col,
    )
    c = ws.cell(TITLE_ROW, 1)
    c.font = Font(bold=True, size=16, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=TITLE_FILL)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def _write_warning(ws, last_col: int, layout_warnings: List[str]) -> None:
    msg = (
        "⚠ このシートは自動生成です。図形はドラッグ/リサイズ/テキスト編集できますが、"
        "再生成時に「フロー図」シートは上書きされます。"
        "手書きメモは「注記」シートに残してください。"
    )
    if layout_warnings:
        msg += " [layout warnings: " + " / ".join(layout_warnings[:3]) + "]"
    ws.cell(WARNING_ROW, 1, msg)
    ws.merge_cells(
        start_row=WARNING_ROW, start_column=1,
        end_row=WARNING_ROW, end_column=last_col,
    )
    c = ws.cell(WARNING_ROW, 1)
    c.font = Font(italic=True, color="806000")
    c.fill = PatternFill("solid", fgColor=WARNING_FILL)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _write_lane_bands(
    ws, flow: Flow, layout: LayoutResult, grid: _Grid,
    last_row: int, last_col: int,
) -> None:
    thin = Side(style="thin", color="8A99B3")
    body_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    medium = Side(style="medium", color="333333")
    header_border = Border(left=medium, right=medium, top=medium, bottom=medium)

    lane_by_id = {l.id: l for l in flow.lanes}

    for i, lane_id in enumerate(layout.lane_order):
        lane = lane_by_id[lane_id]
        fill = PatternFill("solid", fgColor=LANE_BODY_FILLS[i % 2])

        if grid.vertical:
            lane_col = V_FIRST_LANE_COL + i
            # Lane header in row 3
            hdr = ws.cell(V_LANE_HEADER_ROW, lane_col, lane.name)
            hdr.fill = PatternFill("solid", fgColor=LANE_HEADER_FILL)
            hdr.font = Font(bold=True, size=12)
            hdr.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            hdr.border = header_border
            for r in range(V_FIRST_STEP_ROW, last_row + 1):
                cell = ws.cell(r, lane_col)
                cell.fill = fill
                cell.border = body_border
        else:
            lane_row = H_FIRST_LANE_ROW + i
            hdr = ws.cell(lane_row, H_LANE_HEADER_COL, lane.name)
            hdr.fill = PatternFill("solid", fgColor=LANE_HEADER_FILL)
            hdr.font = Font(bold=True, size=12)
            hdr.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            hdr.border = header_border
            for c in range(H_FIRST_STEP_COL, last_col + 1):
                cell = ws.cell(lane_row, c)
                cell.fill = fill
                cell.border = body_border


# --- EMU helpers -----------------------------------------------------------

def _emu_to_cell(emu: int, edges: List[int]) -> Tuple[int, int]:
    for i in range(len(edges) - 1):
        if edges[i] <= emu < edges[i + 1]:
            return i, emu - edges[i]
    return len(edges) - 2, emu - edges[-2]


def _two_cell_anchor(
    x_emu: int, y_emu: int, w_emu: int, h_emu: int, grid: _Grid,
) -> str:
    fc, fco = _emu_to_cell(x_emu, grid.col_edges_emu)
    fr, fro = _emu_to_cell(y_emu, grid.row_edges_emu)
    tc, tco = _emu_to_cell(x_emu + w_emu, grid.col_edges_emu)
    tr, tro = _emu_to_cell(y_emu + h_emu, grid.row_edges_emu)
    return (
        f'<xdr:from><xdr:col>{fc}</xdr:col><xdr:colOff>{fco}</xdr:colOff>'
        f'<xdr:row>{fr}</xdr:row><xdr:rowOff>{fro}</xdr:rowOff></xdr:from>'
        f'<xdr:to><xdr:col>{tc}</xdr:col><xdr:colOff>{tco}</xdr:colOff>'
        f'<xdr:row>{tr}</xdr:row><xdr:rowOff>{tro}</xdr:rowOff></xdr:to>'
    )


# --- drawing XML ------------------------------------------------------------

def _build_drawing_xml(
    flow: Flow, layout: LayoutResult, grid: _Grid,
) -> str:
    ox, oy = grid.body_origin_emu()

    node_shape_id: Dict[str, int] = {}
    next_id = 2
    for nid in layout.nodes:
        node_shape_id[nid] = next_id
        next_id += 1

    node_by_id = {n.id: n for n in flow.nodes}

    parts: List[str] = [
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
        '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"'
        ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"'
        ' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">',
    ]

    label_w_emu = LABEL_W_PX * EMU_PER_PX
    label_h_emu = LABEL_H_PX * EMU_PER_PX

    edge_parts: List[str] = []
    label_parts: List[str] = []

    for e in layout.edges:
        src_nl = layout.nodes[e.from_id]
        dst_nl = layout.nodes[e.to_id]
        src_cx = ox + src_nl.x * EMU_PER_PX
        src_cy = oy + src_nl.y * EMU_PER_PX
        dst_cx = ox + dst_nl.x * EMU_PER_PX
        dst_cy = oy + dst_nl.y * EMU_PER_PX
        w_emu = src_nl.width * EMU_PER_PX
        h_emu = src_nl.height * EMU_PER_PX

        same_lane_back = e.is_back and src_nl.lane_id == dst_nl.lane_id
        if e.is_back:
            src_site, dst_site = _pick_back_connector_sites(
                grid.vertical, same_lane_back, src_nl, dst_nl,
            )
        else:
            src_site, dst_site = _pick_forward_connector_sites(grid.vertical)
        src_pt = _site_point(src_cx, src_cy, w_emu, h_emu, src_site)
        dst_pt = _site_point(dst_cx, dst_cy, w_emu, h_emu, dst_site)

        if e.is_back:
            channel = _back_edge_channel_emu(
                src_nl, dst_nl, src_pt, dst_pt, grid,
            )
            path_pts = _back_edge_points(src_pt, dst_pt, channel, grid.vertical)
        else:
            path_pts = _forward_edge_points(src_pt, dst_pt, grid.vertical)

        edge_parts.append(_connector_path_xml(next_id, path_pts, grid))
        next_id += 1

        if e.condition:
            lx, ly = _label_anchor_emu(
                src_cx, src_cy, dst_cx, dst_cy, w_emu, h_emu,
                label_w_emu, label_h_emu,
                is_back=e.is_back, vertical=grid.vertical,
            )
            label_parts.append(_label_xml(
                next_id, lx, ly, label_w_emu, label_h_emu, e.condition, grid,
            ))
            next_id += 1

    # Draw edges first so node fills cover connector stubs and crossings.
    parts.extend(edge_parts)

    for nid, nl in layout.nodes.items():
        node = node_by_id[nid]
        w_emu = nl.width * EMU_PER_PX
        h_emu = nl.height * EMU_PER_PX
        x_emu = ox + nl.x * EMU_PER_PX - w_emu // 2
        y_emu = oy + nl.y * EMU_PER_PX - h_emu // 2
        parts.append(_node_shape_xml(
            node_shape_id[nid], node, x_emu, y_emu, w_emu, h_emu, grid,
        ))

    parts.extend(label_parts)
    parts.append("</xdr:wsDr>")
    return "\n".join(parts)


def _back_edge_channel_emu(
    src_nl, dst_nl,
    src_pt: Tuple[int, int], dst_pt: Tuple[int, int],
    grid: _Grid,
) -> int:
    """Choose the detour line for a back edge.

    Same-lane returns are local: sending them through the global back
    channel makes a short "send back" action jump to the top of the
    sheet, which is visually noisy. Cross-lane returns use the corridor
    between the source and target lanes.
    """
    gap = LOCAL_BACK_CHANNEL_GAP_PX * EMU_PER_PX
    if grid.vertical:
        if src_nl.lane_id != dst_nl.lane_id:
            left = min(src_pt[0], dst_pt[0])
            right = max(src_pt[0], dst_pt[0])
            return (left + right) // 2
        return max(src_pt[0], dst_pt[0]) + gap

    if src_nl.lane_id != dst_nl.lane_id:
        top = min(src_pt[1], dst_pt[1])
        bottom = max(src_pt[1], dst_pt[1])
        return (top + bottom) // 2

    return max(src_pt[1], dst_pt[1]) + gap


def _back_edge_points(
    src_pt: Tuple[int, int], dst_pt: Tuple[int, int], channel: int, vertical: bool,
) -> List[Tuple[int, int]]:
    sx, sy = src_pt
    dx, dy = dst_pt

    if vertical:
        return [
            (sx, sy),
            (channel, sy),
            (channel, dy),
            (dx, dy),
        ]
    return [
        (sx, sy),
        (sx, channel),
        (dx, channel),
        (dx, dy),
    ]


def _forward_edge_points(
    src_pt: Tuple[int, int], dst_pt: Tuple[int, int], vertical: bool,
) -> List[Tuple[int, int]]:
    sx, sy = src_pt
    dx, dy = dst_pt

    if vertical:
        if abs(dx - sx) < 2:
            return [(sx, sy), (dx, dy)]
        mid_y = (sy + dy) // 2
        return [(sx, sy), (sx, mid_y), (dx, mid_y), (dx, dy)]

    if abs(dy - sy) < 2:
        return [(sx, sy), (dx, dy)]
    mid_x = (sx + dx) // 2
    return [(sx, sy), (mid_x, sy), (mid_x, dy), (dx, dy)]


def _connector_path_xml(
    conn_id: int, points: List[Tuple[int, int]], grid: _Grid,
) -> str:
    """Render an explicit connector path.

    We intentionally avoid ``stCxn``/``endCxn`` here. LibreOffice can
    reinterpret bound connectors during PDF export, especially custom
    U-shaped back edges; fixed paths match the generated image reliably.
    """
    pad = CONNECTOR_PAD_PX * EMU_PER_PX
    xs = [p[0] for p in points]
    ys = [p[1] for p in points]
    bx = max(min(xs) - pad, 0)
    by = max(min(ys) - pad, 0)
    bw = max(max(xs) + pad - bx, 1)
    bh = max(max(ys) + pad - by, 1)

    # Remap to normalised 0..100000 local coordinates.
    def norm(x: int, y: int) -> Tuple[int, int]:
        nx = 0 if bw == 0 else int((x - bx) * 100000 / bw)
        ny = 0 if bh == 0 else int((y - by) * 100000 / bh)
        return nx, ny

    path_xml_parts: List[str] = []
    for i, (px, py) in enumerate(points):
        nx, ny = norm(px, py)
        if i == 0:
            path_xml_parts.append(f'<a:moveTo><a:pt x="{nx}" y="{ny}"/></a:moveTo>')
        else:
            path_xml_parts.append(f'<a:lnTo><a:pt x="{nx}" y="{ny}"/></a:lnTo>')
    path_body = "".join(path_xml_parts)

    anchor = _two_cell_anchor(bx, by, bw, bh, grid)
    return (
        '<xdr:twoCellAnchor editAs="oneCell">'
        f'{anchor}'
        '<xdr:sp macro="" textlink="">'
        f'<xdr:nvSpPr><xdr:cNvPr id="{conn_id}" name="Edge_{conn_id}"/>'
        '<xdr:cNvSpPr/>'
        '</xdr:nvSpPr>'
        '<xdr:spPr>'
        f'<a:xfrm><a:off x="{bx}" y="{by}"/><a:ext cx="{bw}" cy="{bh}"/></a:xfrm>'
        '<a:custGeom>'
        '<a:avLst/><a:gdLst/><a:ahLst/><a:cxnLst/>'
        '<a:rect l="0" t="0" r="100000" b="100000"/>'
        '<a:pathLst>'
        '<a:path w="100000" h="100000">'
        f'{path_body}'
        '</a:path>'
        '</a:pathLst>'
        '</a:custGeom>'
        f'<a:ln w="12700"><a:solidFill><a:srgbClr val="{EDGE_COLOR}"/></a:solidFill>'
        '<a:tailEnd type="triangle"/></a:ln>'
        '</xdr:spPr>'
        '</xdr:sp>'
        '<xdr:clientData/>'
        '</xdr:twoCellAnchor>'
    )


def _site_point(
    cx: int, cy: int, w_emu: int, h_emu: int, site: int,
) -> Tuple[int, int]:
    """Return the absolute EMU coordinate of a shape's connection site."""
    half_w = w_emu // 2
    half_h = h_emu // 2
    if site == SITE_TOP:
        return (cx, cy - half_h)
    if site == SITE_BOTTOM:
        return (cx, cy + half_h)
    if site == SITE_RIGHT:
        return (cx + half_w, cy)
    return (cx - half_w, cy)  # SITE_LEFT


def _pick_forward_connector_sites(vertical: bool) -> Tuple[int, int]:
    """Pick facing connector sites for normal forward edges."""
    if vertical:
        return SITE_BOTTOM, SITE_TOP
    return SITE_RIGHT, SITE_LEFT


def _pick_back_connector_sites(
    vertical: bool, same_lane_back: bool, src_nl, dst_nl,
) -> Tuple[int, int]:
    """Pick connector sites for returns.

    Same-lane returns leave from the outer side of the source so their
    origin is obvious. Cross-lane returns enter the target from the side
    facing the source instead of jumping through the global back channel.
    """
    if vertical:
        if same_lane_back:
            return SITE_TOP, SITE_RIGHT
        if src_nl.x > dst_nl.x:
            return SITE_LEFT, SITE_RIGHT
        return SITE_RIGHT, SITE_LEFT

    if same_lane_back:
        return SITE_RIGHT, SITE_BOTTOM
    if src_nl.y > dst_nl.y:
        return SITE_TOP, SITE_BOTTOM
    return SITE_BOTTOM, SITE_TOP


def _label_anchor_emu(
    src_cx: int, src_cy: int,
    dst_cx: int, dst_cy: int,
    node_w_emu: int, node_h_emu: int,
    label_w_emu: int, label_h_emu: int,
    *, is_back: bool, vertical: bool,
) -> Tuple[int, int]:
    """Place the condition label near the source exit, offset so it
    never sits on the source or target node.

    Rules:
      * back edge (horizontal) — above the source (where the top-channel
        U-turn starts).
      * back edge (vertical)   — to the left of the source.
      * forward (horizontal)   — to the right of the source, shifted
        in y toward the destination so siblings with different dst y
        separate naturally.
      * forward (vertical)     — below the source, shifted in x toward
        the destination for sibling separation.
    """
    half_w = node_w_emu // 2
    half_h = node_h_emu // 2
    lw = label_w_emu // 2
    lh = label_h_emu // 2
    gap = 10 * EMU_PER_PX

    if vertical:
        if is_back:
            cx = src_cx - half_w - gap - lw
            cy = src_cy
        else:
            # Place closer to the destination than the source so labels
            # for stacked sibling targets separate visibly.
            cx = src_cx + (dst_cx - src_cx) * 3 // 4
            cy = src_cy + half_h + gap + lh
    else:
        if is_back:
            cx = src_cx
            cy = src_cy - half_h - gap - lh
        else:
            cx = src_cx + half_w + gap + lw
            # Place closer to the destination than the source so labels
            # for stacked sibling targets separate visibly.
            cy = src_cy + (dst_cy - src_cy) * 3 // 4
    return cx - lw, cy - lh


def _node_shape_xml(
    shape_id: int, node: Node,
    x_emu: int, y_emu: int, w_emu: int, h_emu: int,
    grid: _Grid,
) -> str:
    preset = NODE_PRESET.get(node.type, "flowChartProcess")
    fill = NODE_FILL.get(node.type, "FFFFFF")
    text = xml_escape(node.label)
    name = xml_escape(f"{node.id}_{node.label}")
    anchor = _two_cell_anchor(x_emu, y_emu, w_emu, h_emu, grid)
    return (
        '<xdr:twoCellAnchor editAs="oneCell">'
        f'{anchor}'
        '<xdr:sp macro="" textlink="">'
        f'<xdr:nvSpPr><xdr:cNvPr id="{shape_id}" name="{name}"/><xdr:cNvSpPr/></xdr:nvSpPr>'
        '<xdr:spPr>'
        f'<a:xfrm><a:off x="{x_emu}" y="{y_emu}"/><a:ext cx="{w_emu}" cy="{h_emu}"/></a:xfrm>'
        f'<a:prstGeom prst="{preset}"><a:avLst/></a:prstGeom>'
        f'<a:solidFill><a:srgbClr val="{fill}"/></a:solidFill>'
        '<a:ln w="12700"><a:solidFill><a:srgbClr val="333333"/></a:solidFill></a:ln>'
        '</xdr:spPr>'
        '<xdr:txBody>'
        '<a:bodyPr wrap="square" anchor="ctr"/><a:lstStyle/>'
        '<a:p><a:pPr algn="ctr"/>'
        '<a:r><a:rPr lang="ja-JP" sz="1100" b="1">'
        '<a:solidFill><a:srgbClr val="000000"/></a:solidFill>'
        f'</a:rPr><a:t>{text}</a:t></a:r></a:p>'
        '</xdr:txBody>'
        '</xdr:sp>'
        '<xdr:clientData/>'
        '</xdr:twoCellAnchor>'
    )


def _bent_connector_xml(
    conn_id: int,
    src_shape_id: int, src_site: int,
    dst_shape_id: int, dst_site: int,
    bx: int, by: int, bw: int, bh: int,
    flip_h: bool, flip_v: bool,
    grid: _Grid,
) -> str:
    flips = []
    if flip_h:
        flips.append('flipH="1"')
    if flip_v:
        flips.append('flipV="1"')
    flip_attr = (" " + " ".join(flips)) if flips else ""

    anchor = _two_cell_anchor(bx, by, bw, bh, grid)

    return (
        '<xdr:twoCellAnchor editAs="oneCell">'
        f'{anchor}'
        '<xdr:cxnSp macro="">'
        f'<xdr:nvCxnSpPr><xdr:cNvPr id="{conn_id}" name="Edge_{conn_id}"/>'
        '<xdr:cNvCxnSpPr>'
        f'<a:stCxn id="{src_shape_id}" idx="{src_site}"/>'
        f'<a:endCxn id="{dst_shape_id}" idx="{dst_site}"/>'
        '</xdr:cNvCxnSpPr>'
        '</xdr:nvCxnSpPr>'
        '<xdr:spPr>'
        f'<a:xfrm{flip_attr}><a:off x="{bx}" y="{by}"/><a:ext cx="{bw}" cy="{bh}"/></a:xfrm>'
        '<a:prstGeom prst="bentConnector3"><a:avLst/></a:prstGeom>'
        f'<a:ln w="12700"><a:solidFill><a:srgbClr val="{EDGE_COLOR}"/></a:solidFill>'
        '<a:tailEnd type="triangle"/></a:ln>'
        '</xdr:spPr>'
        '</xdr:cxnSp>'
        '<xdr:clientData/>'
        '</xdr:twoCellAnchor>'
    )


def _label_xml(
    shape_id: int,
    x_emu: int, y_emu: int, w_emu: int, h_emu: int,
    text: str, grid: _Grid,
) -> str:
    t = xml_escape(text)
    anchor = _two_cell_anchor(x_emu, y_emu, w_emu, h_emu, grid)
    return (
        '<xdr:twoCellAnchor editAs="oneCell">'
        f'{anchor}'
        '<xdr:sp macro="" textlink="">'
        f'<xdr:nvSpPr><xdr:cNvPr id="{shape_id}" name="Label_{shape_id}"/>'
        '<xdr:cNvSpPr txBox="1"/></xdr:nvSpPr>'
        '<xdr:spPr>'
        f'<a:xfrm><a:off x="{x_emu}" y="{y_emu}"/><a:ext cx="{w_emu}" cy="{h_emu}"/></a:xfrm>'
        '<a:prstGeom prst="rect"><a:avLst/></a:prstGeom>'
        '<a:solidFill><a:srgbClr val="FFFFFF"/></a:solidFill>'
        '<a:ln><a:noFill/></a:ln>'
        '</xdr:spPr>'
        '<xdr:txBody>'
        '<a:bodyPr wrap="square" anchor="ctr" lIns="18000" tIns="9000" rIns="18000" bIns="9000"/>'
        '<a:lstStyle/>'
        '<a:p><a:pPr algn="ctr"/>'
        f'<a:r><a:rPr lang="ja-JP" sz="1000" b="1">'
        f'<a:solidFill><a:srgbClr val="{LABEL_COLOR}"/></a:solidFill>'
        f'</a:rPr><a:t>{t}</a:t></a:r></a:p>'
        '</xdr:txBody>'
        '</xdr:sp>'
        '<xdr:clientData/>'
        '</xdr:twoCellAnchor>'
    )


# --- xlsx ZIP injection -----------------------------------------------------

DRAWING_CONTENT_TYPE = (
    '<Override PartName="/xl/drawings/drawing1.xml" '
    'ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>'
)
DRAWING_REL_TYPE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"
)
_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def _inject_drawing(xlsx_bytes: bytes, drawing_xml: str) -> bytes:
    in_buf = io.BytesIO(xlsx_bytes)
    out_buf = io.BytesIO()

    with zipfile.ZipFile(in_buf, "r") as zin:
        names = zin.namelist()
        existing_rels = (
            zin.read("xl/worksheets/_rels/sheet1.xml.rels")
            if "xl/worksheets/_rels/sheet1.xml.rels" in names
            else None
        )
        new_rels_bytes, r_id = _build_sheet_rels(existing_rels)

        with zipfile.ZipFile(out_buf, "w", zipfile.ZIP_DEFLATED) as zout:
            for name in names:
                if name == "xl/worksheets/_rels/sheet1.xml.rels":
                    continue
                data = zin.read(name)
                if name == "xl/worksheets/sheet1.xml":
                    data = _add_drawing_ref(data, r_id)
                elif name == "[Content_Types].xml":
                    data = _ensure_drawing_content_type(data)
                zout.writestr(name, data)

            zout.writestr("xl/drawings/drawing1.xml", drawing_xml.encode("utf-8"))
            zout.writestr("xl/worksheets/_rels/sheet1.xml.rels", new_rels_bytes)

    return out_buf.getvalue()


def _build_sheet_rels(existing: bytes | None) -> Tuple[bytes, str]:
    if existing is None:
        xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            f'<Relationship Id="rId1" Type="{DRAWING_REL_TYPE}" Target="../drawings/drawing1.xml"/>'
            "</Relationships>"
        )
        return xml.encode("utf-8"), "rId1"

    text = existing.decode("utf-8")
    max_id = 0
    for m in re.finditer(r'Id="rId(\d+)"', text):
        max_id = max(max_id, int(m.group(1)))
    new_id = f"rId{max_id + 1}"
    new_rel = (
        f'<Relationship Id="{new_id}" Type="{DRAWING_REL_TYPE}" '
        'Target="../drawings/drawing1.xml"/>'
    )
    text = text.replace("</Relationships>", new_rel + "</Relationships>")
    return text.encode("utf-8"), new_id


def _add_drawing_ref(sheet_bytes: bytes, r_id: str) -> bytes:
    text = sheet_bytes.decode("utf-8")
    if "<drawing " in text:
        return sheet_bytes

    if "xmlns:r=" not in text[:500]:
        text = re.sub(
            r"<worksheet\b([^>]*)>",
            lambda m: f'<worksheet{m.group(1)} xmlns:r="{_REL_NS}">',
            text,
            count=1,
        )

    ref = f'<drawing r:id="{r_id}"/>'
    text = text.replace("</worksheet>", ref + "</worksheet>")
    return text.encode("utf-8")


def _ensure_drawing_content_type(ct_bytes: bytes) -> bytes:
    text = ct_bytes.decode("utf-8")
    if "drawings/drawing1.xml" in text:
        return ct_bytes
    text = text.replace("</Types>", DRAWING_CONTENT_TYPE + "</Types>")
    return text.encode("utf-8")
